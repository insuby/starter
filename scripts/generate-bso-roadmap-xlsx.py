#!/usr/bin/env python3
"""Generate print-friendly Excel roadmap for BSO project."""

import datetime
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins


THIN = Side(style='thin', color='CBD5E1')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_HEADER = PatternFill('solid', fgColor='1E40AF')
FILL_SECTION = PatternFill('solid', fgColor='E2E8F0')
FILL_DONE = PatternFill('solid', fgColor='DCFCE7')
FILL_STUB = PatternFill('solid', fgColor='FEF9C3')
FILL_PLANNED = PatternFill('solid', fgColor='F8FAFC')
FILL_SUMMARY = PatternFill('solid', fgColor='EFF6FF')

FONT_TITLE = Font(name='Calibri', size=16, bold=True, color='1E40AF')
FONT_SUBTITLE = Font(name='Calibri', size=11, color='475569')
FONT_HEADER = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
FONT_SECTION = Font(name='Calibri', size=11, bold=True, color='1E293B')
FONT_NORMAL = Font(name='Calibri', size=10, color='1E293B')
FONT_BOLD = Font(name='Calibri', size=10, bold=True, color='1E293B')

STATUS_FILLS = {
    'Сделано': FILL_DONE,
    'Заглушка': FILL_STUB,
    'Запланировано': FILL_PLANNED,
}


ROADMAP = [
    ('1. Фронтенд (интерфейс для пользователей)', None, None, None, None),
    ('1.1', 'Общая оболочка: шапка, меню, переключение кабинетов по ролям', 'Сделано', '—',
     'Оператор центра, округа, ОМУ, ВК, комиссар, аудитор — каждый видит свой интерфейс'),
    ('1.2', 'Дашборд центра: KPI и график операций за 7 дней', 'Сделано', '—',
     'Сводная картина: в обороте, выдано, на удержании, списано'),
    ('1.3', 'Реестр бланков: поиск, фильтры, вкладки «Активные» / «Архив»', 'Сделано', '—',
     'Поиск по номеру (АА123456), фильтр по статусу, типу, месту'),
    ('1.4', 'Карточка бланка с историей всех операций', 'Сделано', '—',
     'Полная «биография» бланка: кто, когда, откуда-куда'),
    ('1.5', 'Мастер «Поступление номеров от типографии» (4 шага)', 'Сделано', '—',
     'Ввод диапазона, проверка серии, разбивка по типам'),
    ('1.6', 'Каскадное распределение по иерархии военкоматов', 'Заглушка', 'Июль',
     'Центр → округ → ОМУ → ВК субъекта → ВК МО; сохранение пока демо'),
    ('1.7', 'Операции с бланками: выдача, возврат, замена, хранение, списание', 'Заглушка', 'Июль',
     'Пошаговые формы готовы; данные пока не сохраняются на сервере'),
    ('1.8', 'Очередь «На подписании» для комиссара', 'Заглушка', 'Август',
     'Список операций, ожидающих подписи; счётчик в шапке'),
    ('1.9', 'Перемещения бланков между ВК', 'Заглушка', 'Июль',
     'Выбор получателя, диапазон номеров, история — без сохранения'),
    ('1.10', 'Форма отчётов (тип, период, уровень)', 'Заглушка', 'Июль–Август',
     'Параметры выбраны, но таблица и Excel — заглушка'),
    ('1.11', 'Раздел «Администрирование»', 'Заглушка', 'Август',
     'Карточки «Пользователи», «Организации», «Настройки» без действий'),
    ('1.12', 'Подключение интерфейса к реальному серверу (вместо демо-данных)', 'Запланировано', 'Июнь–Июль',
     'Все экраны начнут показывать актуальные данные из БД'),
    ('1.13', 'Постраничная загрузка реестра бланков (1500+ ВК МО)', 'Запланировано', 'Июль',
     'Чтобы система не «тормозила» при большом объёме'),
    ('1.14', 'Реальная генерация отчётов: таблица на экране + Excel/PDF', 'Запланировано', 'Июль–Август',
     'Движение, остатки, выданные, история, статистика'),
    ('1.15', 'Рабочая админка: пользователи, организации, справочники', 'Запланировано', 'Август',
     'Управление без обращения к разработчикам'),
    ('1.16', 'Журнал действий пользователей в интерфейсе', 'Запланировано', 'Август',
     'Кто, когда, что изменил — для контроля и аудита'),
    ('1.17', 'Убрать временный переключатель кабинетов, привязать к учётной записи', 'Запланировано', 'Июнь',
     'После входа через домен — автоматически нужный кабинет'),
    ('1.18', 'Обработка ошибок и понятные сообщения при сбоях связи', 'Запланировано', 'Июль',
     '«Сервер недоступен», «нет прав» — без технического жаргона'),
    ('1.19', 'Финальная приёмка интерфейса с заказчиком', 'Запланировано', 'Август',
     'Обход всех сценариев по ролям, фиксация замечаний'),

    ('2. Бэкенд (серверная логика и база данных)', None, None, None, None),
    ('2.1', 'Проектирование базы данных (бланки, операции, оргструктура, пользователи)', 'Запланировано', 'Июнь, нед. 1–2',
     'Основа для хранения всех данных, а не демо в памяти'),
    ('2.2', 'Разработка серверного API по согласованной схеме', 'Запланировано', 'Июнь, нед. 2–4',
     '«Язык» общения между интерфейсом и сервером уже описан'),
    ('2.3', 'Реестр бланков: поиск, фильтры, карточка, история', 'Запланировано', 'Июнь–Июль',
     'Реальные данные вместо демо-списка'),
    ('2.4', 'Поступление серий номеров от типографии', 'Запланировано', 'Июль, нед. 1',
     'Создание записей бланков в БД по диапазону'),
    ('2.5', 'Распределение квот по подразделениям с проверкой лимитов', 'Запланировано', 'Июль, нед. 1–2',
     'Нельзя «раздать» больше, чем поступило'),
    ('2.6', 'Операции: выдача, возврат, замена, хранение, списание', 'Запланировано', 'Июль, нед. 2–3',
     'Статус бланка меняется атомарно, без потерь'),
    ('2.7', 'Очередь на подписание и фиксация результата ЭЦП', 'Запланировано', 'Июль–Август',
     'Операция «зависает» до подписи комиссара'),
    ('2.8', 'Перемещения между ВК с журналом и подтверждением', 'Запланировано', 'Июль, нед. 3',
     'Бланк «в пути» → принят получателем'),
    ('2.9', 'Генерация отчётов на сервере', 'Запланировано', 'Август, нед. 1–2',
     'Тяжёлые выборки — на сервере, не в браузере'),
    ('2.10', 'Справочники: типы бланков, статусы, причины операций', 'Запланировано', 'Июль',
     'Единые правила для всех военкоматов'),
    ('2.11', 'Загрузка реальной оргструктуры (~1500 ВК МО)', 'Запланировано', 'Июль–Август',
     'Сейчас — демо на 1 округ'),
    ('2.12', 'Журнал аудита (все изменения в системе)', 'Запланировано', 'Август',
     'Неизменяемая история для проверок'),
    ('2.13', 'API для внешнего дашборда / аналитики', 'Запланировано', 'Август',
     'Сводные показатели для руководства'),
    ('2.14', 'Резервное копирование и восстановление БД', 'Запланировано', 'Август',
     'Защита от потери данных учёта бланков'),

    ('3. Инфраструктура и развёртывание (закрытая корпоративная сеть)', None, None, None, None),
    ('3.1', 'Выделение отдельного доменного имени для БСО (напр. bso.gomu.local)', 'Запланировано', 'Июнь, нед. 1',
     'Отдельный адрес в корпоративной сети, не смешивается с порталом ГОМУ'),
    ('3.2', 'Согласование с ИБ: зона размещения, порты, доступ из сети', 'Запланировано', 'Июнь, нед. 1–2',
     'Работа только внутри закрытого контура, без выхода в интернет'),
    ('3.3', 'Подготовка серверов: тестовый (staging) и боевой (prod)', 'Запланировано', 'Июнь, нед. 2–3',
     'Сначала проверяем на копии, потом выкатываем «в бой»'),
    ('3.4', 'Контейнеризация (Docker): фронт, бэкенд, БД', 'Запланировано', 'Июнь–Июль',
     'Единый способ установки на Astra Linux'),
    ('3.5', 'Веб-сервер (Apache/Nginx) + прокси к приложению', 'Запланировано', 'Июнь–Июль',
     'Как на портале ГОМУ: Apache на хосте → контейнер'),
    ('3.6', 'TLS-сертификат для домена БСО (внутренний CA)', 'Запланировано', 'Июль',
     'Шифрование трафика внутри сети'),
    ('3.7', 'CI/CD: автоматическая сборка и выкладка на staging', 'Запланировано', 'Июль',
     'Обновления без ручной «сборки на коленке»'),
    ('3.8', 'Мониторинг: доступность, место на диске, ошибки', 'Запланировано', 'Август',
     'О проблеме узнаём до пользователей'),
    ('3.9', 'План аварийного восстановления (RTO/RPO)', 'Запланировано', 'Август',
     'Что делать, если сервер или БД недоступны'),
    ('3.10', 'Документация для эксплуатации (инструкция администратора)', 'Запланировано', 'Август',
     'Передача на сопровождение штатными силами'),
    ('3.11', 'Нагрузочное тестирование (массовые операции, отчёты)', 'Запланировано', 'Август',
     'Проверка перед массовым вводом'),
    ('3.12', 'Перенос на боевой сервер, финальный go-live', 'Запланировано', 'Конец августа',
     'Система доступна пользователям в продуктиве'),

    ('4. Безопасность и доступ (доменный сервер организации)', None, None, None, None),
    ('4.1', 'Вход через домен организации (AD/LDAP / FreeIPA)', 'Запланировано', 'Июнь–Июль',
     'Без отдельных паролей — как на портале ГОМУ'),
    ('4.2', 'Автовход через Kerberos (SSO) в корпоративной сети', 'Запланировано', 'Июль',
     'Открыл браузер на рабочем месте — уже в системе'),
    ('4.3', 'Разграничение прав по 7 ролям (RBAC)', 'Запланировано', 'Июль',
     'Каждый видит только то, что положено по должности'),
    ('4.4', 'Привязка пользователя к организации (центр / округ / ВК МО)', 'Запланировано', 'Июль',
     'Оператор ВК МО не может работать с чужим военкоматом'),
    ('4.5', 'Резервный вход логин/пароль (для админов и аварий)', 'Запланировано', 'Июль',
     'На случай недоступности домена'),
    ('4.6', 'Синхронизация учётных записей домен → БСО', 'Запланировано', 'Июль–Август',
     'Новый сотрудник в домене → появляется в системе'),
    ('4.7', 'Интеграция с криптопровайдером / УКЭП комиссара', 'Запланировано', 'Август',
     'Юридически значимая подпись операций'),
    ('4.8', 'Настройка доверенных URI в браузерах (GPO)', 'Запланировано', 'Июль',
     'Чтобы Kerberos работал на всех рабочих местах'),
    ('4.9', 'Аудит безопасности перед вводом в эксплуатацию', 'Запланировано', 'Август',
     'Проверка ИБ: права, логи, шифрование'),
    ('4.10', 'Политика хранения и удаления персональных данных', 'Запланировано', 'Август',
     'Соответствие требованиям по ФИО граждан'),

    ('5. Интеграции с другими системами', None, None, None, None),
    ('5.1', 'Описание формата JSON-обмена с брокером (v1.0)', 'Сделано', '—',
     'Спецификация готова, пример upr-4_chaes.json'),
    ('5.2', 'Согласование атрибутов обмена с управлениями (upr-1…5)', 'Запланировано', 'Июнь–Июль',
     'Единый «язык» данных между центром и управлениями'),
    ('5.3', 'Отправка данных на брокер', 'Запланировано', 'Июль–Август',
     'Автоматическая выгрузка учёта вверх по иерархии'),
    ('5.4', 'Приём данных от брокера / управлений', 'Запланировано', 'Август',
     'Двусторонний обмен'),
    ('5.5', 'Интеграция со справочником граждан (ЕРН / внутренний реестр)', 'Запланировано', 'Июль–Август',
     'Поиск по ФИО — не демо-записи, а реальная база'),
    ('5.6', 'Связь с организационным справочником военкоматов', 'Запланировано', 'Июль',
     'Актуальная структура без ручного ввода'),
    ('5.7', 'Интеграция с порталом ГОМУ (единый вход, ссылка из меню)', 'Запланировано', 'Август',
     'Пользователь не запоминает второй адрес и пароль'),
    ('5.8', 'Уведомления (опционально): почта / внутренний мессенджер', 'Запланировано', 'Август',
     '«У вас 5 операций на подписании»'),

    ('6. Тестирование, обучение, ввод в эксплуатацию', None, None, None, None),
    ('6.1', 'Тестовый контур (staging) для проверки сценариев', 'Запланировано', 'Июль',
     '«Песочница» без риска для боевых данных'),
    ('6.2', 'Сценарное тестирование по ролям (7 кабинетов)', 'Запланировано', 'Июль–Август',
     'Полный путь: поступление → распределение → выдача → подпись'),
    ('6.3', 'Пилот на 1–2 военкоматах', 'Запланировано', 'Август',
     'Реальная работа ограниченной группы'),
    ('6.4', 'Инструкции пользователя по ролям', 'Запланировано', 'Август',
     'Краткие памятки: оператор ВК МО, комиссар, центр'),
    ('6.5', 'Обучение ключевых пользователей', 'Запланировано', 'Август',
     '1–2 сессии для операторов и комиссаров'),
    ('6.6', 'Устранение замечаний пилота', 'Запланировано', 'Август',
     'Доработки по обратной связи'),
    ('6.7', 'Акт ввода в промышленную эксплуатацию', 'Запланировано', 'Конец августа',
     'Формальное завершение проекта этапа 1'),
    ('6.8', 'План этапа 2 (масштабирование на все управления)', 'Запланировано', 'Сентябрь+',
     'После успешного пилота'),
]

SUMMARY_ROWS = [
    ('Интерфейс (UI-прототип)', 'Готов', '11 разделов, 7 ролей, дизайн из Figma'),
    ('Сервер и база данных', 'Не разработаны', 'Все данные сейчас — демо'),
    ('Авторизация', 'Демо', 'Переключение роли в интерфейсе'),
    ('ЭЦП', 'Демо', 'Имитация подписи комиссара'),
    ('Отчёты', 'Заглушка', 'Форма есть, выгрузка — нет'),
    ('Интеграции', 'Спецификация', 'Формат JSON описан, реализации нет'),
    ('Развёртывание', 'Не начато', 'Нет Docker, CI/CD, выделенного домена'),
]

MONTHLY_ROWS = [
    ('Июнь 2026', 'Фундамент',
     'Домен bso.*, согласование с ИБ, проект БД, старт API, Docker/staging, начало авторизации через домен'),
    ('Июль 2026', 'Ядро системы',
     'Операции с бланками в БД, распределение и перемещения, SSO/Kerberos, справочники, staging готов'),
    ('Август 2026', 'Доводка и запуск',
     'ЭЦП, отчёты Excel/PDF, админка, аудит, нагрузочные тесты, пилот, обучение, go-live'),
]

RISKS = [
    ('Согласование с ИБ и выделение домена', '2–3 недели', 'Без домена не начнётся staging'),
    ('Интеграция ЭЦП', 'Зависит от провайдера', 'Нужны рабочие места комиссаров с УКЭП'),
    ('Справочник граждан', 'Зависит от доступа', 'Без ЕРН/реестра поиск останется ограниченным'),
    ('JSON-обмен с управлениями', 'Согласование формата', 'Может сдвинуть интеграцию с брокером'),
]


def style_cell(cell, font=FONT_NORMAL, fill=None, align='left', wrap=True, border=BORDER):
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(
        horizontal=align,
        vertical='top',
        wrap_text=wrap,
    )
    cell.border = border


def set_print_settings(ws, orientation='landscape'):
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.6, bottom=0.6, header=0.3, footer=0.3)


def build_summary_sheet(wb):
    ws = wb.active
    ws.title = 'Сводка'
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 55

    today = datetime.date.today().strftime('%d.%m.%Y')
    ws.merge_cells('A1:C1')
    c = ws['A1']
    c.value = 'БСО — Учёт бланков строгой отчётности'
    style_cell(c, font=FONT_TITLE, fill=None, align='center', border=Border())
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:C2')
    c = ws['A2']
    c.value = f'Роадмап разработки · План на 3 месяца (июнь–август 2026) · {today}'
    style_cell(c, font=FONT_SUBTITLE, fill=None, align='center', border=Border())
    ws.row_dimensions[2].height = 18

    row = 4
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value='Текущий статус проекта')
    style_cell(c, font=FONT_SECTION, fill=FILL_SECTION, align='left')
    row += 1

    for col, title in enumerate(['Компонент', 'Статус', 'Пояснение'], 1):
        c = ws.cell(row=row, column=col, value=title)
        style_cell(c, font=FONT_HEADER, fill=FILL_HEADER, align='center')
    row += 1

    for comp, status, note in SUMMARY_ROWS:
        ws.cell(row=row, column=1, value=comp)
        ws.cell(row=row, column=2, value=status)
        ws.cell(row=row, column=3, value=note)
        for col in range(1, 4):
            fill = FILL_SUMMARY if status == 'Готов' else FILL_PLANNED
            style_cell(ws.cell(row=row, column=col), fill=fill)
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value='План по месяцам')
    style_cell(c, font=FONT_SECTION, fill=FILL_SECTION, align='left')
    row += 1

    for col, title in enumerate(['Месяц', 'Фокус', 'Ключевые результаты'], 1):
        c = ws.cell(row=row, column=col, value=title)
        style_cell(c, font=FONT_HEADER, fill=FILL_HEADER, align='center')
    row += 1

    for month, focus, results in MONTHLY_ROWS:
        ws.cell(row=row, column=1, value=month)
        ws.cell(row=row, column=2, value=focus)
        ws.cell(row=row, column=3, value=results)
        for col in range(1, 4):
            style_cell(ws.cell(row=row, column=col), fill=FILL_SUMMARY)
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value='Условные обозначения статусов')
    style_cell(c, font=FONT_SECTION, fill=FILL_SECTION, align='left')
    row += 1

    legend = [
        ('Сделано', FILL_DONE, 'Готово и можно показывать'),
        ('Заглушка', FILL_STUB, 'Экран есть, но без реальной работы'),
        ('Запланировано', FILL_PLANNED, 'Предстоит выполнить в указанный срок'),
    ]
    for status, fill, desc in legend:
        ws.cell(row=row, column=1, value=status)
        ws.cell(row=row, column=2, value=desc)
        style_cell(ws.cell(row=row, column=1), font=FONT_BOLD, fill=fill)
        style_cell(ws.cell(row=row, column=2), fill=fill)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value='Основные риски')
    style_cell(c, font=FONT_SECTION, fill=FILL_SECTION, align='left')
    row += 1

    for col, title in enumerate(['Риск', 'Влияние', 'Последствия'], 1):
        c = ws.cell(row=row, column=col, value=title)
        style_cell(c, font=FONT_HEADER, fill=FILL_HEADER, align='center')
    row += 1

    for risk, impact, consequence in RISKS:
        ws.cell(row=row, column=1, value=risk)
        ws.cell(row=row, column=2, value=impact)
        ws.cell(row=row, column=3, value=consequence)
        for col in range(1, 4):
            style_cell(ws.cell(row=row, column=col), fill=FILL_STUB)
        row += 1

    set_print_settings(ws, 'portrait')
    ws.print_title_rows = '1:2'


def build_roadmap_sheet(wb):
    ws = wb.create_sheet('Роадмап')
    widths = [6, 42, 14, 16, 44]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    headers = ['№', 'Задача', 'Статус', 'Срок', 'Комментарий для бизнеса']
    for col, title in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=title)
        style_cell(c, font=FONT_HEADER, fill=FILL_HEADER, align='center')
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    row = 2
    for item in ROADMAP:
        if item[1] is None:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            c = ws.cell(row=row, column=1, value=item[0])
            style_cell(c, font=FONT_SECTION, fill=FILL_SECTION, align='left')
            ws.row_dimensions[row].height = 20
            row += 1
            continue

        num, task, status, deadline, comment = item
        values = [num, task, status, deadline, comment]
        fill = STATUS_FILLS.get(status, FILL_PLANNED)
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            align = 'center' if col in (1, 3, 4) else 'left'
            style_cell(c, fill=fill, align=align)
        ws.row_dimensions[row].height = 36
        row += 1

    set_print_settings(ws, 'landscape')
    ws.print_title_rows = '1:1'


def build_stats_sheet(wb):
    ws = wb.create_sheet('Статистика')
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12

    counts = {'Сделано': 0, 'Заглушка': 0, 'Запланировано': 0}
    sections = {}
    current_section = 'Прочее'

    for item in ROADMAP:
        if item[1] is None:
            current_section = item[0].split('.', 1)[0] + '. ' + item[0].split('. ', 1)[-1][:20]
            sections[current_section] = {'Сделано': 0, 'Заглушка': 0, 'Запланировано': 0}
            continue
        status = item[2]
        counts[status] += 1
        if current_section in sections:
            sections[current_section][status] += 1

    ws['A1'] = 'Сводка по задачам'
    style_cell(ws['A1'], font=FONT_TITLE, border=Border())
    ws.merge_cells('A1:B1')

    row = 3
    for col, title in enumerate(['Статус', 'Кол-во'], 1):
        c = ws.cell(row=row, column=col, value=title)
        style_cell(c, font=FONT_HEADER, fill=FILL_HEADER, align='center')
    row += 1

    total = 0
    for status in ('Сделано', 'Заглушка', 'Запланировано'):
        ws.cell(row=row, column=1, value=status)
        ws.cell(row=row, column=2, value=counts[status])
        for col in (1, 2):
            style_cell(ws.cell(row=row, column=col), fill=STATUS_FILLS[status], align='center' if col == 2 else 'left')
        total += counts[status]
        row += 1

    ws.cell(row=row, column=1, value='Итого')
    ws.cell(row=row, column=2, value=total)
    for col in (1, 2):
        style_cell(ws.cell(row=row, column=col), font=FONT_BOLD, fill=FILL_SECTION, align='center' if col == 2 else 'left')

    row += 2
    ws.cell(row=row, column=1, value='По направлениям')
    style_cell(ws.cell(row=row, column=1), font=FONT_SECTION, fill=FILL_SECTION, border=Border())
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1

    for col, title in enumerate(['Направление', 'Всего задач'], 1):
        c = ws.cell(row=row, column=col, value=title)
        style_cell(c, font=FONT_HEADER, fill=FILL_HEADER, align='center')
    row += 1

    for section, data in sections.items():
        section_total = sum(data.values())
        ws.cell(row=row, column=1, value=section)
        ws.cell(row=row, column=2, value=section_total)
        for col in (1, 2):
            style_cell(ws.cell(row=row, column=col), fill=FILL_SUMMARY, align='center' if col == 2 else 'left')
        row += 1

    set_print_settings(ws, 'portrait')


def main():
    repo = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    out_dir = os.path.join(repo, 'docs')
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, 'Роадмап_БСО.xlsx')

    wb = Workbook()
    build_summary_sheet(wb)
    build_roadmap_sheet(wb)
    build_stats_sheet(wb)
    wb.save(out_path)
    print(out_path)


if __name__ == '__main__':
    main()
